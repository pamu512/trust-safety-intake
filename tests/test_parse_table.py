from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from trust_intake.parse_table import parse_table


def test_csv_totals_and_yoy(tmp_path: Path):
    p = tmp_path / "loss.csv"
    p.write_text(
        "period,brand,loss_eur\n2024-01-01,foodpanda,100\n2025-01-01,foodpanda,128\n",
        encoding="utf-8",
    )
    facts = parse_table(p)
    table = facts["tables"][0]
    assert table["row_count"] == 2
    assert table["totals"]["loss_eur"] == 228
    yoy = next(d for d in facts["derived"] if d["method"] == "yoy")
    assert yoy["value"] == 0.28
    assert yoy["source"] == "csv"


def test_xlsx_multisheet_and_empty_warning(tmp_path: Path):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "loss"
    ws1.append(["period", "loss_eur"])
    ws1.append(["2025-01-01", 50])
    ws2 = wb.create_sheet("empty")
    ws2.append(["period", "loss_eur"])
    path = tmp_path / "t.xlsx"
    wb.save(path)
    facts = parse_table(path)
    names = {t["name"] for t in facts["tables"]}
    assert names == {"loss", "empty"}
    empty = next(t for t in facts["tables"] if t["name"] == "empty")
    assert any("empty" in w.lower() or "single" in w.lower() or "no rows" in w.lower() for w in empty["warnings"])


def test_xlsx_native_date_cells(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "loss"
    ws.append(["period", "loss_eur"])
    ws.append([datetime(2024, 1, 1), 100])
    ws.append([datetime(2025, 1, 1), 128])
    path = tmp_path / "dates.xlsx"
    wb.save(path)
    facts = parse_table(path)
    table = facts["tables"][0]
    period_col = next(c for c in table["columns"] if c["name"] == "period")
    assert period_col["type"] == "date"
    assert table["series"]
    yoy = next(d for d in facts["derived"] if d["method"] == "yoy")
    assert yoy["value"] == 0.28


def test_bad_euro_text_warns(tmp_path: Path):
    p = tmp_path / "bad.csv"
    p.write_text(
        "period,loss_eur\n2024-01-01,100\n2025-01-01,not-a-number\n",
        encoding="utf-8",
    )
    facts = parse_table(p)
    table = facts["tables"][0]
    assert next(c for c in table["columns"] if c["name"] == "loss_eur")["type"] == "number"
    assert table["totals"]["loss_eur"] == 100
    assert any("unparsed number" in w and "not-a-number" in w for w in table["warnings"])
