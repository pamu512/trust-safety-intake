from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

DATE_FMTS = ("%Y-%m-%d", "%Y-%m", "%d/%m/%Y")


def _parse_date(raw: str):
    text = raw.strip()
    for fmt in DATE_FMTS:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _parse_number(raw: str):
    text = raw.strip().replace(",", "").replace("€", "").replace("EUR", "").replace("%", "")
    if text == "":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _looks_numeric(header: str, raw: str) -> bool:
    h = header.lower()
    return any(tok in h or tok in raw for tok in ("€", "eur", "%", "loss", "gmv", "rate"))


def _infer_type(header: str, values: list[str]) -> str:
    nonempty = [v for v in values if v.strip()]
    if nonempty and all(_parse_date(v) for v in nonempty):
        return "date"
    if nonempty and all(_parse_number(v) is not None for v in nonempty):
        return "number"
    if 0 < len(set(nonempty)) <= 12:
        return "category"
    return "text"


def _table_from_rows(name: str, headers: list[str], rows: list[list[str]]) -> tuple[dict, list[dict]]:
    warnings: list[str] = []
    if not rows:
        warnings.append("no rows")
    if len(rows) == 1:
        warnings.append("single row")
    cols = {h: [r[i] if i < len(r) else "" for r in rows] for i, h in enumerate(headers)}
    types = {h: _infer_type(h, cols[h]) for h in headers}
    totals: dict[str, float] = {}
    missingness: dict[str, float] = {}
    n = len(rows)
    for h in headers:
        vals = cols[h]
        miss = 0
        if types[h] == "number":
            nums = []
            for v in vals:
                num = _parse_number(v)
                if num is None:
                    miss += 1
                    if v.strip() or _looks_numeric(h, v):
                        warnings.append(f"unparsed number in {h}: {v!r}")
                else:
                    nums.append(num)
            totals[h] = sum(nums)
        elif types[h] == "date":
            miss = sum(1 for v in vals if not v.strip() or _parse_date(v) is None)
        else:
            miss = sum(1 for v in vals if not v.strip())
        missingness[h] = (miss / n) if n else 0.0
    date_col = next((h for h, t in types.items() if t == "date"), None)
    num_cols = [h for h, t in types.items() if t == "number"]
    series: list[dict] = []
    if date_col and num_cols and n:
        grouped: dict[str, dict[str, float]] = {}
        for r in rows:
            rec = dict(zip(headers, r))
            dt = _parse_date(rec.get(date_col, ""))
            if not dt:
                continue
            period = dt.strftime("%Y-%m-%d")
            bucket = grouped.setdefault(period, {c: 0.0 for c in num_cols})
            for c in num_cols:
                num = _parse_number(rec.get(c, ""))
                if num is not None:
                    bucket[c] += num
        series = [{"period": p, "values": grouped[p]} for p in sorted(grouped)]
    split_col = next((h for h in headers if "brand" in h.lower() or "market" in h.lower()), None)
    splits: list[dict] = []
    if split_col and num_cols:
        grouped_s: dict[str, dict[str, float]] = {}
        for r in rows:
            rec = dict(zip(headers, r))
            key = rec.get(split_col, "") or "unknown"
            bucket = grouped_s.setdefault(key, {c: 0.0 for c in num_cols})
            for c in num_cols:
                num = _parse_number(rec.get(c, ""))
                if num is not None:
                    bucket[c] += num
        splits = [{"key": k, "values": grouped_s[k]} for k in grouped_s]
    derived: list[dict] = []
    for c, total in totals.items():
        derived.append({"name": f"{name}.{c}.sum", "value": total, "unit": None, "source": "csv", "method": "sum"})
    if len(series) >= 2:
        old, new = series[-2], series[-1]
        for c in num_cols:
            ov, nv = old["values"][c], new["values"][c]
            if ov:
                derived.append({"name": f"{name}.{c}.pop", "value": (nv - ov) / ov, "unit": None, "source": "csv", "method": "pop"})
        by_year: dict[int, dict] = {}
        for item in series:
            year = int(item["period"][:4])
            by_year[year] = item
        years = sorted(by_year)
        if len(years) >= 2:
            y0, y1 = years[-2], years[-1]
            if y1 == y0 + 1:
                for c in num_cols:
                    ov, nv = by_year[y0]["values"][c], by_year[y1]["values"][c]
                    if ov:
                        derived.append({"name": f"{name}.{c}.yoy", "value": (nv - ov) / ov, "unit": None, "source": "csv", "method": "yoy"})
    if splits:
        parent = {c: sum(s["values"][c] for s in splits) for c in num_cols}
        for s in splits:
            for c in num_cols:
                if parent[c]:
                    derived.append({"name": f"{name}.{s['key']}.{c}.share", "value": s["values"][c] / parent[c], "unit": None, "source": "csv", "method": "split-share"})
    table = {
        "name": name,
        "columns": [{"name": h, "type": types[h]} for h in headers],
        "row_count": n,
        "totals": totals,
        "missingness": missingness,
        "series": series,
        "splits": splits,
        "warnings": warnings,
    }
    return table, derived


def _read_csv(path: Path) -> list[tuple[str, list[str], list[list[str]]]]:
    with path.open(newline="", encoding="utf-8") as fh:
        reader = csv.reader(fh)
        headers = next(reader, None)
        if not headers:
            return [(path.stem, [], [])]
        rows = [row for row in reader]
        return [(path.stem, headers, rows)]


def _read_xlsx(path: Path) -> list[tuple[str, list[str], list[list[str]]]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    out = []
    for ws in wb.worksheets:
        rows_iter = ws.iter_rows(values_only=True)
        first = next(rows_iter, None)
        if not first:
            out.append((ws.title, [], []))
            continue
        headers = [str(c) if c is not None else "" for c in first]
        rows = []
        for row in rows_iter:
            rows.append(["" if c is None else str(c) for c in row])
        out.append((ws.title, headers, rows))
    return out


def parse_table(path: Path) -> dict:
    if not path.is_file():
        raise FileNotFoundError(path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        sheets = _read_csv(path)
    elif suffix in {".xlsx", ".xlsm"}:
        sheets = _read_xlsx(path)
    else:
        raise ValueError(f"unsupported file type: {suffix}")
    tables = []
    derived: list[dict] = []
    for name, headers, rows in sheets:
        table, der = _table_from_rows(name, headers, rows)
        tables.append(table)
        derived.extend(der)
    return {"tables": tables, "derived": derived}
